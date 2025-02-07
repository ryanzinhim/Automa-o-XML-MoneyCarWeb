import os
import time
import glob
import xml.etree.ElementTree as ET
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import logging
import json
from datetime import datetime, timedelta

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('automacao_xml.log'),
        logging.StreamHandler()
    ]
)

# Configurações
DOWNLOAD_DIR = PASTA
PROGRESS_FILE = os.path.join(DOWNLOAD_DIR, "download_progress.json")

# [Todo o código anterior permanece igual até a definição das funções]

def setup_driver():
    logging.info("Iniciando configuração do Chrome...")
    options = webdriver.ChromeOptions()
    options.add_experimental_option("detach", True)
    download_dir = "C:\\Users\\ryan.alves\\Downloads\\XML"
    
    # Criar diretório de download se não existir
    os.makedirs(download_dir, exist_ok=True)
    
    options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "safebrowsing.enabled": True
    })
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    logging.info("Chrome configurado com sucesso")
    return driver

def login(driver):
    try:
        logging.info("Iniciando processo de login...")
        driver.get("https://www.moneycarweb.com.br/default.aspx")
        
        # Aguardar elementos ficarem visíveis
        wait = WebDriverWait(driver, 10)
        email_field = wait.until(EC.presence_of_element_located((By.ID, "txtEmail")))

        email_field.send_keys(USER)
        
        password_field = driver.find_element(By.ID, "txtSenhaEmail")
        password_field.send_keys(SENHA)
        
        login_button = driver.find_element(By.ID, "Button3")
        login_button.click()
        
        logging.info("Login realizado com sucesso")
        return True
    except Exception as e:
        logging.error(f"Erro no login: {e}")
        return False

def estradanaPag(driver):
    try:
        logging.info("Iniciando pagina download dos XMLs...")
        
        # Clicar no botão Faturamento
        faturamento_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "ctl00_ToolBar1_i7"))
        )
        faturamento_btn.click()
        logging.info("Botão Faturamento clicado")
        
        # Aumentar tempo de espera para garantir carregamento do frame
        time.sleep(7)
        
        # Mudar para o frame com retry
        max_attempts = 3
        for attempt in range(max_attempts):
            try:
                driver.switch_to.frame("master_frame1")
                logging.info("Mudou para o frame correto")
                break
            except Exception as e:
                if attempt == max_attempts - 1:
                    raise
                logging.warning(f"Tentativa {attempt + 1} de mudar para o frame falhou: {e}")
                time.sleep(2)
        
        # Preencher datas com wait explícito
        wait = WebDriverWait(driver, 20)
        campo_data_inicial = wait.until(EC.presence_of_element_located((By.ID, "VNDate1_VNDate1_dtI")))
        campo_data_final = wait.until(EC.presence_of_element_located((By.ID, "VNDate1_VNDate1_dtF")))
        
        driver.execute_script("arguments[0].value = '01/01/2024';", campo_data_inicial)
        driver.execute_script("arguments[0].value = '31/12/2024';", campo_data_final)
        logging.info("Datas preenchidas")
        
        logging.info("Filtro alterado para 50 e evento 'change' disparado")
        driver.execute_script("""
        var selectElement = document.querySelector('[name="VNGridView12$ctl01$ctl10"]');
        selectElement.value = '1000';  // Alterar o valor para 1000
        """)
        logging.info("Filtro alterado para 1000")
        
        # Aguardar botão de pesquisa e clicar
        pesquisar_btn = wait.until(EC.element_to_be_clickable((By.ID, "VNLinkBt1")))
        driver.execute_script("arguments[0].click();", pesquisar_btn)
        logging.info("Pesquisa iniciada")
        
        # Aguardar resultados
        wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(@id, 'VNGridView12_lblDanfe')]")))
        logging.info("Resultados carregados com sucesso")
        
        return True
        
    except Exception as e:
        logging.error(f"Erro ao acessar página de XMLs: {e}")
        # Adicionar screenshot para debug
        try:
            driver.save_screenshot("erro_acesso.png")
            logging.info("Screenshot do erro salvo como erro_acesso.png")
        except:
            pass
        return False

# [O resto do código permanece igual, incluindo as funções download_xmls e main]

def save_progress(current_page, downloaded_count, session_start_time):
    print(f"Tentando salvar progresso em: {os.path.abspath(PROGRESS_FILE)}")  # Adicione esta linha
    progress = {
        "current_page": current_page,
        "downloaded_count": downloaded_count,
        "session_start_time": session_start_time.isoformat()
    }
    with open(PROGRESS_FILE, 'w') as f:
        json.dump(progress, f)

def load_progress():
    try:
        with open(PROGRESS_FILE, 'r') as f:
            progress = json.load(f)
            progress['session_start_time'] = datetime.fromisoformat(progress['session_start_time'])
            return progress
    except FileNotFoundError:
        return {
            "current_page": 1,
            "downloaded_count": 0,
            "session_start_time": datetime.now()
        }

def should_restart_session(session_start_time):
    # Reiniciar 5 minutos antes do timeout de 1 hora
    session_duration = datetime.now() - session_start_time
    return session_duration > timedelta(minutes=55)

def count_existing_xmls():
    """Conta quantos XMLs já existem na pasta de download"""
    return len(glob.glob(os.path.join(DOWNLOAD_DIR, "*.xml")))
def download_xmls(driver, page=1, downloaded_count=0, max_pages=6, session_start_time=None):
    existing_xmls = len(glob.glob(os.path.join(DOWNLOAD_DIR, "*.xml")))
    logging.info(f"XMLs já existentes na pasta: {existing_xmls}")
    
    # Atualizar downloaded_count
    downloaded_count = existing_xmls
    save_progress(page, downloaded_count, session_start_time)
    
    # Inicializar o wait antes de qualquer operação que precise dele
    wait = WebDriverWait(driver, 20)
    try:
        logging.info(f"Página {page}, Downloads feitos: {downloaded_count}")
        
        if page > max_pages:
            logging.info(f"Limite de {max_pages} páginas atingido.")
            return True

        if should_restart_session(session_start_time):
            logging.info("Tempo de sessão próximo do limite. Salvando progresso e reiniciando...")
            save_progress(page, downloaded_count, session_start_time)
            return "RESTART_NEEDED"
        
        # Avançar para a próxima página se já tiver completado os downloads da atual
        if downloaded_count >= page * 1000:
            logging.info(f"Página {page} concluída. Avançando para a próxima página...")
            next_button = driver.find_element(By.XPATH, "//input[@src='images/next2.gif']")
            driver.execute_script("arguments[0].click();", next_button)

            time.sleep(20)
            page += 1  # Incrementa o número da página diretamente

            # Aguardar a nova página carregar
            wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(@id, 'VNGridView12_lblDanfe')]")))

            logging.info(f"Página {page} carregada. Continuando os downloads...")
        time.sleep(10)
        logging.info(f"Iniciando download dos XMLs da página {page}...")
        time.sleep(10)

        xml_elements = wait.until(EC.presence_of_all_elements_located(
            (By.XPATH, "//span[contains(@id, 'VNGridView12_lblDanfe')]")
        ))
        total_elements = len(xml_elements)
        
        # Começar do início da página
        for i in range(total_elements):
            try:
                if should_restart_session(session_start_time):
                    save_progress(page, downloaded_count, session_start_time)
                    return "RESTART_NEEDED"
                
                driver.switch_to.default_content()
                driver.switch_to.frame("master_frame1")

                xml_xpath = f"(//span[contains(@id, 'VNGridView12_lblDanfe')])[{i + 1}]//a[contains(@title, 'Download do XML')]"
                xml_link = wait.until(EC.element_to_be_clickable((By.XPATH, xml_xpath)))
                
                driver.execute_script("arguments[0].scrollIntoView(true);", xml_link)
                time.sleep(1)
                
                before_files = set(os.listdir(DOWNLOAD_DIR))
                driver.execute_script("arguments[0].click();", xml_link)
                
                max_wait = 30
                download_successful = False
                while max_wait > 0:
                    current_files = set(os.listdir(DOWNLOAD_DIR))
                    new_files = current_files - before_files
                    if new_files:
                        downloaded_count += 1
                        logging.info(f"Download XML {downloaded_count} concluído (Página {page}, Item {i + 1}/{total_elements})")
                        time.sleep(2)
                        download_successful = True
                        break
                    time.sleep(1)
                    max_wait -= 1
                
                if not download_successful:
                    logging.warning(f"Timeout esperando download do XML {i + 1} na página {page}")
            
            except Exception as e:
                logging.error(f"Erro ao baixar XML {i + 1} na página {page}: {e}")
                continue
        
        try:
            # Verificar se deve avançar para a próxima página
            next_button = driver.find_element(By.XPATH, "//input[@src='images/next2.gif']")
            if next_button.is_enabled():
                logging.info(f"Página {page} concluída. Avançando...")
                driver.execute_script("arguments[0].click();", next_button)
                time.sleep(23)
                return download_xmls(driver, page + 1, downloaded_count, max_pages, session_start_time)
            else:
                logging.info(f"Processo concluído. Total: {downloaded_count} XMLs.")
                return True
                
        except Exception as e:
            if "no such element" in str(e).lower():
                logging.info(f"Processo concluído. Total: {downloaded_count} XMLs.")
                return True
            else:
                logging.error(f"Erro ao verificar próxima página: {e}")
                return False
                
    except Exception as e:
        logging.error(f"Erro durante o download dos XMLs na página {page}: {e}")
        save_progress(page, downloaded_count, session_start_time)
        return False

                
    except Exception as e:
        logging.error(f"Erro durante o download dos XMLs na página {page}: {e}")
        save_progress(page, downloaded_count, session_start_time)
        return False
def process_xmls():
    try:
        logging.info("Iniciando processamento dos XMLs...")
        download_path = "C:\\Users\\ryan.alves\\Downloads\\XML"
        excel_path = "C:\\Users\\ryan.alves\\Downloads\\Sem título 1.xlsx"
        
        wb = load_workbook(excel_path)
        ws = wb["Número e chave de acesso - Mone"]
        
        # Definir cabeçalhos
        ws['A1'] = "Numero"
        ws['B1'] = "Chave"
        row = 2
        
        xml_files = glob.glob(os.path.join(download_path, "*.xml"))
        total_files = len(xml_files)
        
        processed_files = 0
        for xml_file in xml_files:
            try:
                tree = ET.parse(xml_file)
                root = tree.getroot()
                ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe'}
                
                nNF = root.find(".//nfe:nNF", ns).text if root.find(".//nfe:nNF", ns) is not None else "ERRO"
                chave = root.find(".//nfe:infNFe", ns).attrib.get('Id', "")[3:] if root.find(".//nfe:infNFe", ns) is not None else "ERRO"
                
                ws[f"A{row}"] = nNF
                ws[f"B{row}"] = chave
                row += 1
                processed_files += 1
                logging.info(f"Processado XML {processed_files}/{total_files}: NF {nNF}")
                logging.info(f"Processado XML {processed_files}/{total_files}: id {chave}")

            except Exception as e:
                logging.error(f"Erro ao processar {xml_file}: {e}")
        
        wb.save(excel_path)
        logging.info(f"Processo concluído. {processed_files} arquivos processados com sucesso.")
        return True
    except Exception as e:
        logging.error(f"Erro no processamento dos XMLs: {e}")
        return False


def main():
    while True:
        try:
            # Carregar progresso atual
            progress = load_progress()
            current_page = progress["current_page"]
            downloaded_count = progress["downloaded_count"]
            session_start_time = datetime.now()
            
            # Verificar quantos XMLs já existem
            existing_xmls = count_existing_xmls()
            logging.info(f"Encontrados {existing_xmls} XMLs já baixados na pasta")
            logging.info(f"Continuando de: Página {current_page}, Downloads {downloaded_count}")
            
            driver = setup_driver()
            
            if not login(driver):
                raise Exception("Falha no login")
            
            if not estradanaPag(driver):
                raise Exception("Falha ao acessar página")
            
            # Se não estiver na primeira página, navegar até a página atual
            if current_page > 1:
                for _ in range(current_page - 1):
                    next_button = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, "//input[@src='images/next2.gif']"))
                    )
                    driver.execute_script("arguments[0].click();", next_button)
                    time.sleep(5)
            
            # Iniciar/continuar downloads
            result = download_xmls(driver, current_page, downloaded_count, max_pages=6, 
                                 session_start_time=session_start_time)
            
            if result == "RESTART_NEEDED":
                logging.info("Reiniciando sessão...")
                driver.quit()
                time.sleep(5)
                continue
            elif result:
                logging.info("Download completo!")
                break
            else:
                logging.error("Erro durante o download, tentando reiniciar...")
                time.sleep(5)
        
        except Exception as e:
            logging.error(f"Erro na execução: {e}")
            time.sleep(5)
        finally:
            try:
                driver.quit()
            except:
                pass
        # if not process_xmls():
        #     raise Exception("Falha no processamento dos XMLs")
        
        logging.info("Automação concluída com sucesso!")
if __name__ == "__main__":
    main()
##lembrar de trocar a variavel por env.